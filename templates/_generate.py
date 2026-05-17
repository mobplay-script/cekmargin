"""Generator file template gratis: Template Profit per Produk.
Jalankan: python templates/_generate.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

TEAL = "006565"
TEAL_DARK = "004F4F"
INPUT_FILL = "FFF8E1"      # kuning lembut: kolom yang diisi user
HEAD_FILL = TEAL
ROUND_ROWS = 30            # jumlah baris produk

wb = Workbook()

# ---------- Sheet 1: Cara Pakai ----------
cara = wb.active
cara.title = "Cara Pakai"
cara.sheet_view.showGridLines = False

cara["A1"] = "Template Profit per Produk"
cara["A1"].font = Font(name="Arial", size=18, bold=True, color=TEAL)
cara["A2"] = "Toolkit Seller Indonesia - cekmargin.pages.dev"
cara["A2"].font = Font(name="Arial", size=10, color="666666")

panduan = [
    "",
    "Cara memakai template ini:",
    "1. Buka tab 'Profit per Produk' di bawah.",
    "2. Isi hanya kolom berwarna kuning: Nama Produk, Modal, Biaya Packing,",
    "   Biaya Iklan, Harga Jual, Diskon, Fee Marketplace, dan Komisi Affiliate.",
    "3. Kolom putih (Untung Bersih, Margin, Status) terisi otomatis.",
    "4. Lihat kolom Status untuk tahu apakah produk Aman, Tipis, Berisiko, atau Rugi.",
    "",
    "Catatan penting:",
    "- Isi Diskon, Fee, dan Komisi dalam persen (contoh: ketik 8% atau 0,08).",
    "- Fee marketplace berbeda tiap kategori dan bisa berubah - cek tarif terbaru toko Anda.",
    "- Semua hasil adalah estimasi, bukan jaminan keuntungan.",
    "",
    "Arti Status:",
    "- Aman      : margin >= 15%, cukup sehat untuk menutup retur & biaya tak terduga.",
    "- Tipis     : margin 5%-15%, masih untung tapi hati-hati.",
    "- Berisiko  : margin 0%-5%, satu retur bisa membuat rugi.",
    "- Rugi      : margin <= 0%, harga jual belum menutup biaya.",
    "",
    "Butuh kalkulator cepat? Kunjungi cekmargin.pages.dev",
]
for i, teks in enumerate(panduan, start=4):
    c = cara.cell(row=i, column=1, value=teks)
    bold = teks.endswith(":")
    c.font = Font(name="Arial", size=11, bold=bold,
                  color=TEAL_DARK if bold else "333333")
cara.column_dimensions["A"].width = 90

# ---------- Sheet 2: Profit per Produk ----------
ws = wb.create_sheet("Profit per Produk")
ws.sheet_view.showGridLines = False

headers = [
    ("Nama Produk", 22, True),
    ("Modal (Rp)", 13, True),
    ("Biaya Packing (Rp)", 15, True),
    ("Biaya Iklan/Produk (Rp)", 17, True),
    ("Harga Jual (Rp)", 14, True),
    ("Diskon (%)", 11, True),
    ("Fee Marketplace (%)", 15, True),
    ("Komisi Affiliate (%)", 15, True),
    ("Harga Setelah Diskon (Rp)", 18, False),
    ("Potongan Marketplace (Rp)", 18, False),
    ("Untung Bersih (Rp)", 15, False),
    ("Margin (%)", 11, False),
    ("Status", 14, False),
]

# Judul
ws.merge_cells("A1:M1")
ws["A1"] = "Template Profit per Produk"
ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", start_color=TEAL)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:M2")
ws["A2"] = "Isi kolom kuning. Kolom putih terhitung otomatis. Persen ditulis seperti 8%."
ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")

HEAD_ROW = 4
FIRST = HEAD_ROW + 1
LAST = HEAD_ROW + ROUND_ROWS

thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, (judul, lebar, _is_input) in enumerate(headers, start=1):
    c = ws.cell(row=HEAD_ROW, column=col, value=judul)
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=HEAD_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws.column_dimensions[c.column_letter].width = lebar
ws.row_dimensions[HEAD_ROW].height = 32

rupiah = 'Rp#,##0;-Rp#,##0;"-"'
persen = "0.0%"

contoh = [
    ["Contoh: Kaos Polos", 20000, 2000, 1000, 30000, 0.0, 0.08, 0.0],
    ["Contoh: Tumbler", 35000, 3000, 0, 45000, 0.20, 0.08, 0.0],
]

for r in range(FIRST, LAST + 1):
    idx = r - FIRST
    isi = contoh[idx] if idx < len(contoh) else [None] * 8

    for col in range(1, 9):
        c = ws.cell(row=r, column=col, value=isi[col - 1])
        c.fill = PatternFill("solid", start_color=INPUT_FILL)
        c.border = border
        c.font = Font(name="Arial", size=10, color="0000FF")
        if col in (2, 3, 4, 5):
            c.number_format = rupiah
        elif col in (6, 7, 8):
            c.number_format = persen

    # I: Harga setelah diskon
    ws.cell(row=r, column=9, value=f'=IF(E{r}="","",E{r}*(1-F{r}))')
    # J: Potongan marketplace
    ws.cell(row=r, column=10, value=f'=IF(E{r}="","",I{r}*(G{r}+H{r}))')
    # K: Untung bersih
    ws.cell(row=r, column=11,
            value=f'=IF(E{r}="","",I{r}-J{r}-B{r}-C{r}-D{r})')
    # L: Margin
    ws.cell(row=r, column=12,
            value=f'=IF(OR(E{r}="",I{r}=0),"",K{r}/I{r})')
    # M: Status
    ws.cell(row=r, column=13,
            value=(f'=IF(E{r}="","",IF(L{r}>=0.15,"Aman",'
                   f'IF(L{r}>=0.05,"Tipis",IF(L{r}>0,"Berisiko rugi","Rugi"))))'))

    for col in range(9, 14):
        c = ws.cell(row=r, column=col)
        c.border = border
        c.font = Font(name="Arial", size=10, color="000000")
        if col in (9, 10, 11):
            c.number_format = rupiah
        elif col == 12:
            c.number_format = persen
        if col == 13:
            c.font = Font(name="Arial", size=10, bold=True, color="000000")
            c.alignment = Alignment(horizontal="center")

status_range = f"M{FIRST}:M{LAST}"
ws.conditional_formatting.add(status_range, CellIsRule(
    operator="equal", formula=['"Aman"'],
    fill=PatternFill("solid", start_color="C8E6C9"),
    font=Font(name="Arial", size=10, bold=True, color="137333")))
ws.conditional_formatting.add(status_range, CellIsRule(
    operator="equal", formula=['"Tipis"'],
    fill=PatternFill("solid", start_color="FFF0C2"),
    font=Font(name="Arial", size=10, bold=True, color="9A6700")))
ws.conditional_formatting.add(status_range, CellIsRule(
    operator="equal", formula=['"Berisiko rugi"'],
    fill=PatternFill("solid", start_color="FFE0C2"),
    font=Font(name="Arial", size=10, bold=True, color="9A4A00")))
ws.conditional_formatting.add(status_range, CellIsRule(
    operator="equal", formula=['"Rugi"'],
    fill=PatternFill("solid", start_color="FFD6D2"),
    font=Font(name="Arial", size=10, bold=True, color="93000A")))

ws.freeze_panes = "A5"

out = os.path.join(os.path.dirname(__file__), "Template-Profit-per-Produk.xlsx")
wb.save(out)
print("Tersimpan:", out)
